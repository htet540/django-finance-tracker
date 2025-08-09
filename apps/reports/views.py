import io
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone

from apps.reports.forms import ReportFilterForm
from apps.transactions.models import Transaction

def apply_filters(request):
    qs = (
        Transaction.objects
        .filter(is_deleted=False, entity__is_deleted=False)
        .select_related("entity", "currency", "designated_purpose")
    )
    form = ReportFilterForm(request.GET or None)
    if form.is_valid():
        cd = form.cleaned_data

        # Entity filters
        if cd.get("custom_id"):
            qs = qs.filter(entity__custom_id__iexact=cd["custom_id"].strip())
        if cd.get("entity_name"):
            qs = qs.filter(entity__name__icontains=cd["entity_name"])
        if cd.get("entity_type") in ("donor", "recipient"):
            qs = qs.filter(entity__type=cd["entity_type"])
        if cd.get("location"):
            qs = qs.filter(entity__location__icontains=cd["location"])

        # Tx filters
        if cd.get("currency"):
            qs = qs.filter(currency=cd["currency"])
        if cd.get("date_from"):
            qs = qs.filter(date__gte=cd["date_from"])
        if cd.get("date_to"):
            qs = qs.filter(date__lte=cd["date_to"])
        if cd.get("min_amount") is not None:
            qs = qs.filter(amount__gte=cd["min_amount"])
        if cd.get("max_amount") is not None:
            qs = qs.filter(amount__lte=cd["max_amount"])
        if cd.get("min_converted_mmk") is not None:
            qs = qs.filter(converted_amount_mmk__gte=cd["min_converted_mmk"])
        if cd.get("max_converted_mmk") is not None:
            qs = qs.filter(converted_amount_mmk__lte=cd["max_converted_mmk"])

        order_by = cd.get("order_by") or "-date"
        qs = qs.order_by(order_by)

    totals = qs.aggregate(
        total_amount=Sum("amount"),
        total_mmk=Sum("converted_amount_mmk"),
    )
    totals["total_amount"] = totals["total_amount"] or Decimal("0")
    totals["total_mmk"] = totals["total_mmk"] or Decimal("0")

    return form, qs, totals

@login_required
def index(request):
    form, qs, totals = apply_filters(request)
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get("page") or 1)
    return render(request, "reports/index.html", {
        "form": form,
        "page_obj": page_obj,
        "totals": totals,
        "count": qs.count(),
        "params": request.GET,
    })

@login_required
def export_csv(request):
    _, qs, totals = apply_filters(request)
    import csv
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=transactions_report.csv"
    writer = csv.writer(response)
    headers = ["Date","Entity ID","Entity Name","Type","Currency","Amount","Rate","MMK","Purpose","Location"]
    writer.writerow(headers)
    for t in qs.iterator():
        writer.writerow([
            t.date.strftime("%d/%m/%Y"),
            t.entity.custom_id, t.entity.name, t.entity.get_type_display(),
            t.currency.code, f"{t.amount}", f"{t.exchange_rate}", f"{t.converted_amount_mmk}",
            t.designated_purpose.name if t.designated_purpose else "", t.entity.location or "",
        ])
    writer.writerow([])
    writer.writerow(["","","","","","TOTAL MMK","",f"{totals['total_mmk']}"])
    return response

@login_required
def export_xlsx(request):
    _, qs, totals = apply_filters(request)
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active; ws.title = "Report"
    headers = ["Date","Entity ID","Entity Name","Type","Currency","Amount","Rate","MMK","Purpose","Location"]
    ws.append(headers)
    for t in qs.iterator():
        ws.append([
            t.date.strftime("%d/%m/%Y"), t.entity.custom_id, t.entity.name, t.entity.get_type_display(),
            t.currency.code, float(t.amount), float(t.exchange_rate), float(t.converted_amount_mmk),
            t.designated_purpose.name if t.designated_purpose else "", t.entity.location or "",
        ])
    ws.append([]); ws.append(["","","","","","TOTAL MMK","", float(totals["total_mmk"])])
    for i,_ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    out = io.BytesIO(); wb.save(out); out.seek(0)
    resp = HttpResponse(out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = "attachment; filename=transactions_report.xlsx"
    return resp

@login_required
def export_pdf(request):
    _, qs, totals = apply_filters(request)
    landscape = request.GET.get("landscape") in ("1", "true", "yes")
    html = render_to_string("reports/pdf.html", {
        "rows": qs,
        "totals": totals,
        "filters": request.GET,
        "generated_at": timezone.localtime(),
        "landscape": landscape,
    })
    from xhtml2pdf import pisa
    result = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=result, encoding="UTF-8")
    pdf = result.getvalue()
    resp = HttpResponse(pdf, content_type="application/pdf")
    filename = "transactions_report_landscape.pdf" if landscape else "transactions_report.pdf"
    resp["Content-Disposition"] = f"attachment; filename={filename}"
    return resp